from selenium import webdriver # webdriver를 이용해 해당 브라우저를 열기 위해
from selenium.webdriver import ActionChains # 일련의 작업들을(ex.아이디 입력, 비밀번호 입력, 로그인 버튼 클릭...) 연속적으로 실행할 수 있게 하기 위해
from selenium.webdriver.common.keys import Keys # 키보드 입력을 할 수 있게 하기 위해
from selenium.webdriver.common.by import By # html요소 탐색을 할 수 있게 하기 위해
from selenium.webdriver.support.ui import WebDriverWait # 브라우저의 응답을 기다릴 수 있게 하기 위해
from selenium.webdriver.support import expected_conditions as EC # html요소의 상태를 체크할 수 있게 하기 위해
# 이 외에도 필요한 모듈이 있다면 따로 호출해준다.
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
from openpyxl import load_workbook
import time

# var
year = '2024'
file = 'bojo_srv.xlsx'

wb = load_workbook(file, read_only=False, data_only=False)

# 브라우저 꺼짐 방지 옵션
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options) #크롬 열기
action = ActionChains(driver)

driver.set_window_size(1524,768)

driver.get('https://www.bojo.go.kr/bojo.do') # 해당 웹사이트로 접속
time.sleep(2)

for srv in range(7,18):
    

    driver.refresh()

    driver.find_element(By.XPATH,'//*[@id="_notice2"]/section/header/button').click() # pop-up
    time.sleep(2)

    action.move_to_element(driver.find_element(By.XPATH,'//*[@id="skip_gnb"]/div/ul/li[2]/a')).perform() # 페이지 이동
    time.sleep(2)
    driver.find_element(By.XPATH, '//*[@id="skip_gnb"]/div/ul/li[2]/div/div/ul/li[1]/a').click()
    time.sleep(2)

    year_select = Select(driver.find_element(By.ID, "selAsstnSe")) # 사업구분 opt
    year_select.select_by_value('1')
    time.sleep(2)


    # driver.find_elements(By.XPATH, f'//*[@id="skin_content"]/section/article/div/div[2]/div/div[4]/div[2]/div/div[1]/div/li[{srv+1}]/label/input').click() #서비스 주제 opt
    driver.find_element(By.XPATH, f'/html/body/div[2]/section/section/article/div/div[2]/div/div[4]/div[2]/div/div[1]/div/ul/li[{srv+1}]/label/input').click()
    
    iter = 1
    skip_opt = False

    srv_label = driver.find_element(By.XPATH, f'/html/body/div[2]/section/section/article/div/div[2]/div/div[4]/div[2]/div/div[1]/div/ul/li[{srv+1}]')
    srv_text = srv_label.text
    
    srv_text = srv_text.replace('/', '')

    # print(srv_text)

    if srv_text not in wb.sheetnames:
        wb.create_sheet(srv_text)
    Xtable = wb[srv_text]  # 해당 시트를 가져오기
    wb.save(filename=file) # xlsx save

    if Xtable['A1'].value == 'done': # 끝낸 opt
        continue
    
    driver.find_element(By.XPATH,'//*[@id="skin_content"]/section/article/div/div[2]/div/button').click() # 조회
    time.sleep(2)     

    num_select=Select(driver.find_element(By.ID, "selViewCount1")) # '노출개수' opt
    num_select.select_by_value('50')

    time.sleep(2)

    total = driver.find_element(By.XPATH,'//*[@id="viewDtlbzListTotcnt"]').text #관서별 공시 갯수
    print(srv_text, total)

    
    
    #정보공시 대상 순회
    page = 1 # 페이지 초기화      
    tableWrap = driver.find_elements(By.XPATH, '//*[@id="cardItem1"]//li') # 대상 리스트 들
    while len(tableWrap) != 0:
        print(len(tableWrap))

        #50 목록 순회
        for num, j in enumerate(tableWrap):
            
            # try: # 이미 있는 데이터 건너띄기 50개
            #     find = f'//*[@id="cardItem1"]/li[{50}]/div[2]/div[1]/h5/div/a'
            #     title = driver.find_element(By.XPATH, find)
            #     if Xtable[f'E{iter+50}'].value != '' and (Xtable[f'E{iter+50}'].value in title.text):
            #         print("already, skippp 50")
            #         continue
            # except:
            #     pass
            
            iter+=1
            
            #에외
            if Xtable[f'A{iter}'].value == 'error':
                continue

            # try: # 이미 있는 데이터 건너띄기
            #     find = '//*[@id="tableWrap"]/ul/li['+str(num+1)+']/div[2]/div[1]/h5/div/a'
            #     title = driver.find_element(By.XPATH, find)
            #     print(title.text, Xtable[f'E{num+1}'].value)
            #     if Xtable[f'E{iter}'].value != '' and (Xtable[f'E{iter}'].value in title.text):
            #         print("already, skippp")
            #         continue
            # except:
            #     pass

            Xtable[f'A{iter}'] = iter - 1 # No.   
            Xtable[f'B{iter}'] = driver.find_element(By.XPATH,f'//*[@id="cardItem1"]/li[{num+1}]/div[2]/div[1]/h5/span').text # 중앙관서            
            Xtable[f'C{iter}'] = srv_text #서비스 주제
            Xtable[f'D{iter}'] = driver.find_element(By.XPATH,f'//*[@id="cardItem1"]/li[{num+1}]/div[2]/div[1]/h5/div/a').text.split('. ', 1)[1] #사업명
            # print(driver.find_element(By.XPATH,f'//*[@id="cardItem1"]/li[{num+1}]/div[2]/div[1]/dl').text[5:])
            Xtable[f'E{iter}'] = driver.find_element(By.XPATH,f'//*[@id="cardItem1"]/li[{num+1}]/div[2]/div[1]/dl').text[5:] #사업기간               
            Xtable[f'F{iter}'] = driver.find_element(By.XPATH,f'//*[@id="cardItem1"]/li[{num+1}]/div[2]/div[1]/p').text #사업설명
            wb.save(filename=file) # xlsx save

        #페이지 넘기기 or 넘어가기
        if len(tableWrap) < 50: # 마지막 페이지
            print("brreak")
            break
        time.sleep(2) #StaleElementReferenceException!
        page += 1
        
        if driver.find_element(By.XPATH,f'//*[@id="afterSearch"]/div/div[1]/div[4]/nav/div[1]/button[{page}]').get_attribute('class') == "page next":
            driver.find_element(By.XPATH,f'//*[@id="afterSearch"]/div/div[1]/div[4]/nav/div[1]/button[{page}]').click()
            page = 3
        else:
            driver.find_element(By.XPATH,f'//*[@id="afterSearch"]/div/div[1]/div[4]/nav/div[1]/button[{page}]').click()
       
        time.sleep(2)
        tableWrap = driver.find_elements(By.XPATH, '//*[@id="cardItem1"]//li') # 대상 리스트 업데이트

        wb.save(filename=file) # xlsx save
    wb.save(filename=file) # xlsx save
print("end")
