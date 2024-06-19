from selenium import webdriver # webdriver를 이용해 해당 브라우저를 열기 위해
from selenium.webdriver import ActionChains # 일련의 작업들을(ex.아이디 입력, 비밀번호 입력, 로그인 버튼 클릭...) 연속적으로 실행할 수 있게 하기 위해
from selenium.webdriver.common.keys import Keys # 키보드 입력을 할 수 있게 하기 위해
from selenium.webdriver.common.by import By # html요소 탐색을 할 수 있게 하기 위해
from selenium.webdriver.support.ui import WebDriverWait # 브라우저의 응답을 기다릴 수 있게 하기 위해
from selenium.webdriver.support import expected_conditions as EC # html요소의 상태를 체크할 수 있게 하기 위해
# 이 외에도 필요한 모듈이 있다면 따로 호출해준다.
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
from openpyxl import load_workbook
import time

# var
year = '2022'
file = 'bojo2022.xlsx'
done = 38

# 브라우저 꺼짐 방지 옵션
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options) #크롬 열기
action = ActionChains(driver)

driver.get('https://www.bojo.go.kr/bojo.do') # 해당 웹사이트로 접속

driver.set_window_size(1524,768)
time.sleep(2)


driver.find_element(By.XPATH,'//*[@id="_notice2"]/section/header/button').click() # pop-up
time.sleep(2)

action.move_to_element(driver.find_element(By.XPATH,'//*[@id="skip_gnb"]/div/ul/li[4]/a')).perform() # 페이지 이동
time.sleep(2)
driver.find_element(By.XPATH, '//*[@id="skip_gnb"]/div/ul/li[4]/div/ul/li[1]/ul/li[2]/a').click()
time.sleep(2)

wb = load_workbook(file, read_only=False, data_only=False)
# sheet_names = wb.get_sheet_names()

year_select = Select(driver.find_element(By.ID, "EA001201Frm_bsnsyear")) # 사업연도 opt
year_select.select_by_value(year)
year_select2 = Select(driver.find_element(By.ID, "EA001201Frm_fsyr")) # 회계연도 opt    
year_select2.select_by_value(year)
time.sleep(2)

#목록 크롤링
opts = driver.find_elements(By.XPATH, '//*[@id="EA001201Frm_jrsdCode"]//option')
time.sleep(2)
select=Select(driver.find_element(By.ID, "EA001201Frm_jrsdCode")) # '중앙관서' opt
iter = 1
skip_opt = False
for i, n in enumerate(opts): 
    if i < done:
        continue
    
    iter = 1
    wb.save(filename=file) # xlsx save
    if i==0: # 빈 속성
        continue
    opt_value = n.get_attribute('value')
    select.select_by_value(opt_value)
    try:
        disabled_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="skin_content"]/section/article/div/div[2]/div/button'))
        )
        disabled_button.click()
    except:
        print("요소를 클릭할 수 없습니다.")
    time.sleep(2)
    total = driver.find_element(By.XPATH,'//*[@id="total_num"]').text #관서별 공시 갯수
    print(n.text, total, opt_value)

    driver.find_element(By.XPATH,'//*[@id="skin_content"]/section/article/div/div[2]/div/button').click() # 조회
    time.sleep(2)                 

    if total != '0':
        if n.text not in wb.sheetnames:
            wb.create_sheet(n.text)
        Xtable = wb[n.text]  # 해당 시트를 가져오기
    
    if i==1:
        num_select=Select(driver.find_element(By.ID, "sbxGridPerPageSize")) # '노출개수' opt
        num_select.select_by_value('50')
        skip_opt = False
    
    #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    if i == done:
        num_select=Select(driver.find_element(By.ID, "sbxGridPerPageSize")) # '노출개수' opt
        num_select.select_by_value('50')
        skip_opt = False

    if Xtable['A1'].value == 'done': # 끝낸 opt
        continue
    

    #정보공시 대상 순회
    page = 1 # 페이지 초기화
    tableWrap = driver.find_elements(By.XPATH, '//*[@id="tableWrap"]/ul//li') # 대상 리스트 들
    dp = n.text #중앙관서 이름
    while len(tableWrap) != 0:
        info_button = driver.find_elements(By.XPATH, '//*[@id="tableWrap"]/ul//li/div[2]/div[2]/div/button')
        print(len(info_button))                      #//*[@id="tableWrap"]/ul/li[11]/div[2]/div[2]/div

        

        #50 목록 순회
        for num, j in enumerate(info_button):
            if num == 0: #!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                try: # 이미 있는 데이터 건너띄기
                    find = '//*[@id="tableWrap"]/ul/li['+str(50)+']/div[2]/div[1]/h5/div/a'
                    title = driver.find_element(By.XPATH, find)
                    if Xtable[f'E{iter+50}'].value != '' and (Xtable[f'E{iter+50}'].value in title.text):
                        print("already, skip50")
                        iter+=50
                        break
                except:
                    pass
            
            
            iter+=1
            if skip_opt:
                break

            if iter<4300: #half
                continue
            
            #에외
            if Xtable[f'A{iter}'].value == 'error':
                continue

            #if opt_value == '019' and ((iter > 562 and iter < 566) or (iter > 1252 and iter < 1255) or (iter > 1937 and iter < 1941)): # 2022 고용노동부
            #     continue
            if opt_value == '162' and ((iter > 535 and iter < 538)): # 2022 과기부
                continue
            try: #팝업 ok?
                disabled_button = WebDriverWait(driver, 10).until( 
                        EC.element_to_be_clickable((By.XPATH, f'//*[@id="tableWrap"]/ul/li[{num+1}]/div[2]/div[2]/div/button'))
                )
            except:
                Xtable[f'B{iter}'] = "NOT POPUP" # SKIP 기록
                continue                


            #날짜 설정 
            find = '//*[@id="tableWrap"]/ul/li['+str(num+1)+']/div[2]/div[1]/h5/dl/dd'
            update = driver.find_element(By.XPATH, find).text.split('.')
            print(str(num+1), int(update[0]) , int(update[1]) , int(update[2]))
            if int(update[0]) >2024:
                if int(update[1]) > 4:
                    print("latest, skippp")
                    iter-=1
                elif int(update[1]) == 4:
                    if int(update[2]) > 8:
                        print("latest, skippp")
                        iter-=1

            try: # 이미 있는 데이터 건너띄기
                find = '//*[@id="tableWrap"]/ul/li['+str(num+1)+']/div[2]/div[1]/h5/div/a'
                title = driver.find_element(By.XPATH, find)
                print(title.text, Xtable[f'E{num+1}'].value)
                if Xtable[f'E{iter}'].value != '' and (Xtable[f'E{iter}'].value in title.text):
                    print("already, skippp")
                    continue
            except:
                pass
            
            # j.click() # pop-up 실행
            try:
                disabled_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f'//*[@id="tableWrap"]/ul/li[{num+1}]/div[2]/div[2]/div'))
                )
                disabled_button.click()
                time.sleep(2)
            except:
                Xtable[f'A{iter}'] = 'j.click error'
                continue

            #가짜버튼
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="mainDiv"]'))
                )
            except TimeoutException as e:
                Xtable[f'B{iter}'] = "SKIP" # SKIP 기록
                continue
           
            #보조사업자공시정보
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="displayDiv"]/div/ul/li/div'))
                )
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="mainDiv"]/div[2]/div/div[1]'))
                )
            except TimeoutException as e:
                print(iter,"시간 초과: 요소를 찾을 수 없습니다.<1>")
            try:
                Xtable[f'A{iter}'] = iter - 1 # No.   
                Xtable[f'B{iter}'] = dp # 중앙관서            
                Xtable[f'C{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_fiscalyear"]').text #회계연도
                Xtable[f'D{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_bsnsyear"]').text #사업연도
                Xtable[f'E{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_dtlbzNm"]').text #사업명
                Xtable[f'F{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_excInsttNm2"]').text #기관명
                Xtable[f'G{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_rprsntvNm"]').text #대표자명
                Xtable[f'H{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_telno"]').text #전화번호
                Xtable[f'I{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_rdnmDetailAdres"]').text #소재지
            except:
                Xtable[f'B{iter}'] = "SKIP1" # SKIP 기록
                continue

            #총사업비
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="displayTotMoney1"]'))
                )
            except TimeoutException as e:
                print(iter,"시간 초과: 요소를 찾을 수 없습니다.<2>")
            try:
                Xtable[f'J{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_totGsamt"]').text #총사업비
                Xtable[f'K{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_gsamt"]').text #국고보조금
                Xtable[f'L{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_lsamt"]').text #지자체부담금
                Xtable[f'M{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_salm"]').text #자기부담금
            except:
                Xtable[f'B{iter}'] = "SKIP2" # SKIP 기록
                continue

            #교부신청정보
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="tableGridHtml"]'))
                )
            except TimeoutException as e:
                print(iter,"시간 초과: 요소를 찾을 수 없습니다.<3>")
            try:
                #예산액
                Xtable[f'N{iter}'] = driver.find_element(By.XPATH,'//*[@id="tableGridHtml"]/div/div[2]').text #교부신청액
                Xtable[f'O{iter}'] = driver.find_element(By.XPATH,'//*[@id="tableGridHtml"]/div/div[3]/div/div[1]/div[2]').text #국고보조금
                Xtable[f'P{iter}'] = driver.find_element(By.XPATH,'//*[@id="tableGridHtml"]/div/div[3]/div/div[2]/div[2]').text #지자체부담금
                #교부결정액
                Xtable[f'Q{iter}'] = driver.find_element(By.XPATH,'//*[@id="tableGridHtml"]/div/div[4]/div/div[1]/div[2]').text #신청교부
                Xtable[f'R{iter}'] = driver.find_element(By.XPATH,'//*[@id="tableGridHtml"]/div/div[4]/div/div[2]/div[2]').text #직권교부
            except:
                Xtable[f'B{iter}'] = "SKIP3" # SKIP 기록
                continue

            #사업계획서
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="mainDiv"]/div[2]/div/div[4]/div/table/tbody'))
                )
            except TimeoutException as e:
                print(iter,"시간 초과: 요소를 찾을 수 없습니다.<4>")
            try:
                Xtable[f'S{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_bsnsPurpsCn"]').text #사업목적
                Xtable[f'T{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_bsnsCn"]').text #사업내용
                Xtable[f'U{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_bsnsBeginDe"]').text #사업기간
                Xtable[f'V{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_bsnsPlaceNm"]').text #사업장소
                Xtable[f'W{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_trgterCoCn"]').text #대상자수                
                Xtable[f'X{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_expensBndCn"]').text #보조금 이외의 경비부담내용
                Xtable[f'Y{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_ernAmountProcessMthCn"]').text #수익금액의 처리방법
                Xtable[f'Z{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_assetsDebtMatter"]').text #신청자의 자산/부채
                Xtable[f'AA{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_expcEffectCn"]').text #기대효과
                Xtable[f'AB{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_rsltGoalCn"]').text #성과목표
                Xtable[f'AC{iter}'] = driver.find_element(By.XPATH,'//*[@id="EA001301ViewVal_cnsdrMatter"]').text #고려사항
            except:
                Xtable[f'B{iter}'] = "SKIP4" # SKIP 기록
                continue
            wb.save(filename=file) # xlsx save
            # except NoSuchElementException as e:
            #     print(iter, "요소를 찾을 수 없습니다.")

            #공시정보 pop-up 닫기
            try:
                disabled_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="_popUpCloseBt"]'))
                )
                disabled_button.click()
                time.sleep(2)
            except:
                #재진입
                driver.get('https://www.bojo.go.kr/bojo.do') # 해당 웹사이트로 접속
                time.sleep(2)
                driver.find_element(By.XPATH,'//*[@id="_notice2"]/section/header/button').click() # pop-up
                time.sleep(2)
                action.move_to_element(driver.find_element(By.XPATH,'//*[@id="skip_gnb"]/div/ul/li[4]/a')).perform() # 페이지 이동
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="skip_gnb"]/div/ul/li[4]/div/ul/li[1]/ul/li[2]/a').click()
                time.sleep(2)
                year_select = Select(driver.find_element(By.ID, "EA001201Frm_bsnsyear")) # 사업연도 opt
                year_select.select_by_value(year)
                year_select2 = Select(driver.find_element(By.ID, "EA001201Frm_fsyr")) # 회계연도 opt    
                year_select2.select_by_value(year)
                time.sleep(2)
                skip_opt = True

                Xtable[f'A{iter-1}'] = 'error' # No.
                Xtable[f'A{iter}'] = 'error' # No.

                print(iter, "요소를 클릭할 수 없습니다.")
            
            # driver.find_element(By.XPATH, '//*[@id="_popUpCloseBt"]').click()

        #페이지 넘기기 or 넘어가기
        if len(tableWrap) < 10: # 마지막 페이지 !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            print("brreak")
            break
        time.sleep(2) #StaleElementReferenceException!
        page += 1
        try:
            if driver.find_element(By.XPATH,f'//*[@id="skin_content"]/section/article/div/div[6]/nav/div[1]/button[{page}]').get_attribute('class') == "page next":
                driver.find_element(By.XPATH,f'//*[@id="skin_content"]/section/article/div/div[6]/nav/div[1]/button[{page}]').click()
                page = 3
            else:
                driver.find_element(By.XPATH,f'//*[@id="skin_content"]/section/article/div/div[6]/nav/div[1]/button[{page}]').click()
        except ElementClickInterceptedException as e: # 클릭에 방해
            try:
                #팝업 닫기
                disabled_button = WebDriverWait(driver, 10).until( 
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="_popUpCloseBt"]'))
                )
                disabled_button.click()
            except:
                #재진입
                driver.get('https://www.bojo.go.kr/bojo.do') # 해당 웹사이트로 접속
                time.sleep(2)
                driver.find_element(By.XPATH,'//*[@id="_notice2"]/section/header/button').click() # pop-up
                time.sleep(2)
                action.move_to_element(driver.find_element(By.XPATH,'//*[@id="skip_gnb"]/div/ul/li[4]/a')).perform() # 페이지 이동
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="skip_gnb"]/div/ul/li[4]/div/ul/li[1]/ul/li[2]/a').click()
                time.sleep(2)
                year_select = Select(driver.find_element(By.ID, "EA001201Frm_bsnsyear")) # 사업연도 opt
                year_select.select_by_value(year)
                year_select2 = Select(driver.find_element(By.ID, "EA001201Frm_fsyr")) # 회계연도 opt    
                year_select2.select_by_value(year)
                time.sleep(2)
                skip_opt = True

                Xtable[f'A{iter-1}'] = 'error' # No.
                Xtable[f'A{iter}'] = 'error' # No.

                print(iter, "요소를 클릭할 수 없습니다.")
        time.sleep(2)
        tableWrap = driver.find_elements(By.XPATH, '//*[@id="tableWrap"]/ul//li') # 대상 리스트 업데이트
        wb.save(filename=file) # xlsx save
    wb.save(filename=file) # xlsx save
print("end")
